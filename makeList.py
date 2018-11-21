# inputされるファイルはUTF-8でエンコーディングされているものじゃないと動かない

import csv
import openpyxl
import zenhan

with open('sample.csv', 'r') as f:
    reader = csv.reader(f)
    count = 0
    wb = openpyxl.Workbook()
    ws = wb.active
    row_n = 2

    # ヘッダー生成
    ws['A1'] = '学籍番号'
    ws['B1'] = '学生氏名'
    ws['C1'] = 'カナ氏名'
    ws['D1'] = '学年'
    ws['E1'] = 'コース'
    ws['F1'] = '第3回'
    ws['G1'] = '第4回'
    ws['H1'] = '第5回'
    ws['I1'] = '第6回'
    ws['J1'] = '第7回'
    ws['K1'] = '第8回'
    ws['L1'] = '合計'

    for row in reader:
        if count < 3:  # 最初の3行の情報は不要なため削除
            count += 1
            continue

        ws['A' + str(row_n)] = int(row[0])  # 学籍番号
        ws['B' + str(row_n)] = row[1]  # 学生氏名
        ws['C' + str(row_n)] = zenhan.h2z(row[2])  # カナ氏名
        ws['D' + str(row_n)] = int(row[4])  # 学年
        ws['E' + str(row_n)] = row[5][8:]  # コース
        row_n += 1

# 最後のスクリプトの定義
ws['A' + str(row_n)] = '出席人数'

ws['F' + str(row_n)] = '=COUNTIF(F2:F' + str(row_n - 1) + ',"○")'
ws['G' + str(row_n)] = '=COUNTIF(G2:G' + str(row_n - 1) + ',"○")'
ws['H' + str(row_n)] = '=COUNTIF(H2:H' + str(row_n - 1) + ',"○")'
ws['I' + str(row_n)] = '=COUNTIF(I2:I' + str(row_n - 1) + ',"○")'
ws['J' + str(row_n)] = '=COUNTIF(J2:J' + str(row_n - 1) + ',"○")'
ws['K' + str(row_n)] = '=COUNTIF(K2:K' + str(row_n - 1) + ',"○")'
ws['L' + str(row_n)] = '=AVERAGE(F' + str(row_n) + ':K' + str(row_n) + ')'

wb.save('sample.xlsx')